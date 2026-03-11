#!/usr/bin/env python3
"""
Script de génération d'un fichier Excel exemple pour l'import de missions GesRes
"""

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from datetime import datetime
    
    # Créer un nouveau classeur
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Missions"
    
    # En-têtes avec style
    headers = [
        'Numéro de mission',
        'Date début',
        'Date fin',
        'Description',
        'Titre',
        'Code brigade',
        'Effectifs requis',
        'Priorité',
        'Compétences'
    ]
    
    # Style d'en-tête
    header_fill = PatternFill(start_color="0070C0", end_color="0070C0", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    
    # Écrire les en-têtes
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Données exemples (missions BTA PERSAN de mars 2026)
    data = [
        ['1819992', '13/03/26 15:00:00', '13/03/26 23:00:00', 'Ordre et sécurité publique', 'Renfort PAM', '58577', 2, 'normale', 'PSC1'],
        ['1827585', '14/03/26 00:30:00', '14/03/26 07:00:00', 'Ordre et sécurité publique', 'Renfort BGE', 'BTA-PERSAN', 1, 'haute', ''],
        ['1819702', '14/03/26 15:00:00', '14/03/26 23:00:00', 'Ordre et sécurité publique', 'Renfort PAM', '58577', 2, 'normale', ''],
        ['1827585', '15/03/26 00:30:00', '15/03/26 07:00:00', 'Ordre et sécurité publique', 'Renfort BGE', '58577', 1, 'normale', ''],
        ['1820123', '15/03/26 15:00:00', '15/03/26 23:00:00', 'Ordre et sécurité publique', 'Renfort PAM', '58577', 2, 'normale', 'PSC1, Permis B']
    ]
    
    # Écrire les données
    for row_idx, row_data in enumerate(data, start=2):
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(horizontal='left', vertical='center')
    
    # Ajuster la largeur des colonnes
    column_widths = [18, 20, 20, 30, 20, 15, 18, 12, 25]
    for col, width in enumerate(column_widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width
    
    # Ajouter une note explicative
    ws.cell(row=8, column=1, value="Note : Remplacez ces données d'exemple par vos propres missions.")
    ws.cell(row=8, column=1).font = Font(italic=True, color="666666", size=9)
    
    ws.cell(row=9, column=1, value="Format des dates : JJ/MM/AA HH:MM:SS ou DD/MM/YYYY HH:MM:SS")
    ws.cell(row=9, column=1).font = Font(italic=True, color="666666", size=9)
    
    ws.cell(row=10, column=1, value="Priorités possibles : basse, normale, haute, urgente")
    ws.cell(row=10, column=1).font = Font(italic=True, color="666666", size=9)
    
    # Sauvegarder
    wb.save('/home/user/webapp/modele_import_missions.xlsx')
    print("✅ Fichier Excel créé : /home/user/webapp/modele_import_missions.xlsx")
    
except ImportError:
    print("⚠️ openpyxl n'est pas installé. Installation...")
    import subprocess
    subprocess.run(['pip', 'install', 'openpyxl'], check=True)
    print("✅ openpyxl installé. Relancer le script.")
